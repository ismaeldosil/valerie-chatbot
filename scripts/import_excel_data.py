#!/usr/bin/env python3
"""Import supplier data from Excel file into SQLite database.

This script reads PO history data from an Excel file and aggregates it into
the Valerie chatbot database schema.

Usage:
    python scripts/import_excel_data.py --excel-path PATH --db-path data/valerie.db
"""

import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import typer
from openpyxl import load_workbook
from rich.console import Console
from rich.progress import (
    BarColumn,
    Progress,
    SpinnerColumn,
    TaskProgressColumn,
    TextColumn,
    TimeElapsedColumn,
    TimeRemainingColumn,
)
from sqlalchemy import select
from sqlalchemy.orm import Session

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from valerie.data.database import Database
from valerie.data.schema import (
    Category,
    LegalEntity,
    Supplier,
    SupplierCategory,
    SupplierItem,
)

app = typer.Typer(help="Import supplier data from Excel into SQLite database.")
console = Console()


# Column mapping from Excel headers to internal names
COLUMN_MAP = {
    "Sold-to Legal Entity": "legal_entity",
    "Purchase Order Number": "po_number",
    "Buyer": "buyer",
    "Supplier Name": "supplier_name",
    "Supplier Site": "supplier_site",
    "Open Date": "creation_date",
    "Purchase Order Line Number": "line_number",
    "Item": "item_code",
    "Item Description": "item_description",
    "Category Name": "category",
    "UOM": "uom",
    "Net Ordered Quantity": "quantity",
    "Purchase Price": "unit_price",
    "Ordered Amount": "amount",
    "Concatenated Segments": "segments",
    "Supplier Item": "supplier_item_code",
    "Long Description": "long_description",
}


@dataclass
class SupplierAgg:
    """Aggregation data for a supplier."""

    name: str
    site: str = ""
    total_orders: int = 0
    total_amount: float = 0.0
    first_order_date: Optional[datetime] = None
    last_order_date: Optional[datetime] = None
    po_numbers: set = field(default_factory=set)


@dataclass
class ItemAgg:
    """Aggregation data for a supplier item."""

    item_code: str
    supplier_name: str
    description: str = ""
    supplier_item_code: str = ""
    category: str = ""
    uom: str = "EA"
    prices: list = field(default_factory=list)
    total_qty: float = 0.0
    total_amount: float = 0.0
    order_count: int = 0
    last_order_date: Optional[datetime] = None


@dataclass
class CategoryAgg:
    """Aggregation data for a category."""

    name: str
    level1: str = ""
    level2: str = ""
    level3: str = ""
    item_count: int = 0
    total_amount: float = 0.0


@dataclass
class SupplierCategoryAgg:
    """Aggregation data for supplier-category junction."""

    supplier_name: str
    category_name: str
    item_count: int = 0
    total_amount: float = 0.0


@dataclass
class LegalEntityAgg:
    """Aggregation data for a legal entity."""

    name: str
    total_orders: int = 0
    total_amount: float = 0.0
    po_numbers: set = field(default_factory=set)


def parse_category(category_str: str) -> tuple[str, str, str]:
    """Parse category string into level1, level2, level3.

    Category format examples:
    - "Controlled Material-Chemicals-Acetone" -> ("Controlled Material", "Chemicals", "Acetone")
    - "Non-Controlled Service-WWT Chemicals/Supplies-Monthly Service"
       -> ("Non-Controlled Service", "WWT Chemicals/Supplies", "Monthly Service")
    - "CIP-CIP-Outside Services" -> ("CIP", "CIP", "Outside Services")

    The first part indicates if controlled/non-controlled and material/service type.
    The second part is the subcategory.
    The third part (and beyond) is the specific item type.

    Args:
        category_str: The category string to parse.

    Returns:
        Tuple of (level1, level2, level3).
    """
    if not category_str:
        return ("", "", "")

    # Handle special prefix patterns that should not be split
    # Pattern: "Non-Controlled" or "Controlled" followed by "Material" or "Service"
    prefixes = [
        "Non-Controlled Material",
        "Non-Controlled Service",
        "Controlled Material",
        "Controlled Service",
    ]

    level1 = ""
    remaining = category_str

    for prefix in prefixes:
        if category_str.startswith(prefix + "-"):
            level1 = prefix
            remaining = category_str[len(prefix) + 1:]  # +1 for the hyphen
            break

    if not level1:
        # No special prefix found, use simple split
        parts = category_str.split("-")
        level1 = parts[0].strip() if len(parts) > 0 else ""
        level2 = parts[1].strip() if len(parts) > 1 else ""
        level3 = "-".join(parts[2:]).strip() if len(parts) > 2 else ""
        return (level1, level2, level3)

    # Split remaining into level2 and level3
    parts = remaining.split("-")
    level2 = parts[0].strip() if len(parts) > 0 else ""
    level3 = "-".join(parts[1:]).strip() if len(parts) > 1 else ""

    return (level1, level2, level3)


def read_excel_rows(excel_path: Path, progress: Progress):
    """Read rows from Excel file with progress tracking.

    Args:
        excel_path: Path to the Excel file.
        progress: Rich Progress instance.

    Yields:
        Dict of column name -> value for each row.
    """
    wb = load_workbook(str(excel_path), read_only=True, data_only=True)
    sheet = wb.active

    # Get headers from first row
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    # Map headers to internal names
    header_map = {}
    for i, header in enumerate(headers):
        if header in COLUMN_MAP:
            header_map[i] = COLUMN_MAP[header]

    # Count total rows (approximate for progress)
    # For read-only mode, we can't easily get max_row, so estimate
    task = progress.add_task("[cyan]Reading Excel rows...", total=None)

    row_count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, value in enumerate(row):
            if i in header_map:
                row_dict[header_map[i]] = value
        yield row_dict
        row_count += 1
        if row_count % 1000 == 0:
            progress.update(task, advance=1000, description=f"[cyan]Read {row_count:,} rows...")

    progress.update(task, completed=row_count, total=row_count)
    wb.close()


def aggregate_data(excel_path: Path, progress: Progress):
    """Aggregate data from Excel rows.

    Args:
        excel_path: Path to the Excel file.
        progress: Rich Progress instance.

    Returns:
        Tuple of (suppliers, items, categories, supplier_categories, legal_entities).
    """
    suppliers: dict[str, SupplierAgg] = {}
    items: dict[tuple[str, str], ItemAgg] = {}  # (supplier_name, item_code)
    categories: dict[str, CategoryAgg] = {}
    supplier_categories: dict[tuple[str, str], SupplierCategoryAgg] = {}
    legal_entities: dict[str, LegalEntityAgg] = {}

    for row in read_excel_rows(excel_path, progress):
        supplier_name = str(row.get("supplier_name") or "").strip()
        if not supplier_name:
            continue

        supplier_site = str(row.get("supplier_site") or "").strip()
        item_code = str(row.get("item_code") or "").strip()
        item_description = str(row.get("item_description") or "").strip()
        supplier_item_code = str(row.get("supplier_item_code") or "").strip()
        category_name = str(row.get("category") or "").strip()
        uom = str(row.get("uom") or "EA").strip()
        unit_price = float(row.get("unit_price") or 0)
        quantity = float(row.get("quantity") or 0)
        amount = float(row.get("amount") or 0)
        creation_date = row.get("creation_date")
        po_number = str(row.get("po_number") or "").strip()
        legal_entity_name = str(row.get("legal_entity") or "").strip()

        # Convert creation_date to datetime if needed
        if creation_date and not isinstance(creation_date, datetime):
            try:
                creation_date = datetime.fromisoformat(str(creation_date))
            except ValueError:
                creation_date = None

        # Aggregate supplier data
        if supplier_name not in suppliers:
            suppliers[supplier_name] = SupplierAgg(name=supplier_name, site=supplier_site)

        sup = suppliers[supplier_name]
        sup.total_amount += amount
        if po_number and po_number not in sup.po_numbers:
            sup.po_numbers.add(po_number)
            sup.total_orders += 1

        if creation_date:
            if sup.first_order_date is None or creation_date < sup.first_order_date:
                sup.first_order_date = creation_date
            if sup.last_order_date is None or creation_date > sup.last_order_date:
                sup.last_order_date = creation_date

        # Aggregate item data
        if item_code:
            item_key = (supplier_name, item_code)
            if item_key not in items:
                items[item_key] = ItemAgg(
                    item_code=item_code,
                    supplier_name=supplier_name,
                    description=item_description,
                    supplier_item_code=supplier_item_code,
                    category=category_name,
                    uom=uom,
                )

            item = items[item_key]
            if unit_price > 0:
                item.prices.append(unit_price)
            item.total_qty += quantity
            item.total_amount += amount
            item.order_count += 1
            if creation_date:
                if item.last_order_date is None or creation_date > item.last_order_date:
                    item.last_order_date = creation_date

        # Aggregate category data
        if category_name:
            if category_name not in categories:
                level1, level2, level3 = parse_category(category_name)
                categories[category_name] = CategoryAgg(
                    name=category_name,
                    level1=level1,
                    level2=level2,
                    level3=level3,
                )

            cat = categories[category_name]
            cat.item_count += 1
            cat.total_amount += amount

            # Aggregate supplier-category data
            sc_key = (supplier_name, category_name)
            if sc_key not in supplier_categories:
                supplier_categories[sc_key] = SupplierCategoryAgg(
                    supplier_name=supplier_name,
                    category_name=category_name,
                )

            sc = supplier_categories[sc_key]
            sc.item_count += 1
            sc.total_amount += amount

        # Aggregate legal entity data
        if legal_entity_name:
            if legal_entity_name not in legal_entities:
                legal_entities[legal_entity_name] = LegalEntityAgg(name=legal_entity_name)

            le = legal_entities[legal_entity_name]
            le.total_amount += amount
            if po_number and po_number not in le.po_numbers:
                le.po_numbers.add(po_number)
                le.total_orders += 1

    return suppliers, items, categories, supplier_categories, legal_entities


def insert_or_update_data(
    session: Session,
    suppliers: dict[str, SupplierAgg],
    items: dict[tuple[str, str], ItemAgg],
    categories: dict[str, CategoryAgg],
    supplier_categories: dict[tuple[str, str], SupplierCategoryAgg],
    legal_entities: dict[str, LegalEntityAgg],
    progress: Progress,
):
    """Insert or update aggregated data into the database.

    Args:
        session: SQLAlchemy session.
        suppliers: Aggregated supplier data.
        items: Aggregated item data.
        categories: Aggregated category data.
        supplier_categories: Aggregated supplier-category data.
        legal_entities: Aggregated legal entity data.
        progress: Rich Progress instance.
    """
    # Insert/update categories first (needed for foreign keys)
    task = progress.add_task("[green]Inserting categories...", total=len(categories))
    category_id_map: dict[str, int] = {}

    for cat_agg in categories.values():
        existing = session.execute(
            select(Category).where(Category.name == cat_agg.name)
        ).scalar_one_or_none()

        if existing:
            existing.level1 = cat_agg.level1
            existing.level2 = cat_agg.level2
            existing.level3 = cat_agg.level3
            existing.item_count = cat_agg.item_count
            existing.total_amount = cat_agg.total_amount
            category_id_map[cat_agg.name] = existing.id
        else:
            cat = Category(
                name=cat_agg.name,
                level1=cat_agg.level1,
                level2=cat_agg.level2,
                level3=cat_agg.level3,
                item_count=cat_agg.item_count,
                total_amount=cat_agg.total_amount,
            )
            session.add(cat)
            session.flush()
            category_id_map[cat_agg.name] = cat.id

        progress.advance(task)

    session.commit()

    # Insert/update suppliers
    task = progress.add_task("[green]Inserting suppliers...", total=len(suppliers))
    supplier_id_map: dict[str, int] = {}

    for sup_agg in suppliers.values():
        avg_order_value = sup_agg.total_amount / sup_agg.total_orders if sup_agg.total_orders > 0 else 0

        existing = session.execute(
            select(Supplier).where(Supplier.name == sup_agg.name)
        ).scalar_one_or_none()

        if existing:
            existing.site = sup_agg.site
            existing.total_orders = sup_agg.total_orders
            existing.total_amount = sup_agg.total_amount
            existing.avg_order_value = avg_order_value
            existing.first_order_date = sup_agg.first_order_date
            existing.last_order_date = sup_agg.last_order_date
            supplier_id_map[sup_agg.name] = existing.id
        else:
            sup = Supplier(
                name=sup_agg.name,
                site=sup_agg.site,
                total_orders=sup_agg.total_orders,
                total_amount=sup_agg.total_amount,
                avg_order_value=avg_order_value,
                first_order_date=sup_agg.first_order_date,
                last_order_date=sup_agg.last_order_date,
            )
            session.add(sup)
            session.flush()
            supplier_id_map[sup_agg.name] = sup.id

        progress.advance(task)

    session.commit()

    # Insert/update legal entities
    task = progress.add_task("[green]Inserting legal entities...", total=len(legal_entities))

    for le_agg in legal_entities.values():
        existing = session.execute(
            select(LegalEntity).where(LegalEntity.name == le_agg.name)
        ).scalar_one_or_none()

        if existing:
            existing.total_orders = le_agg.total_orders
            existing.total_amount = le_agg.total_amount
        else:
            le = LegalEntity(
                name=le_agg.name,
                total_orders=le_agg.total_orders,
                total_amount=le_agg.total_amount,
            )
            session.add(le)

        progress.advance(task)

    session.commit()

    # Insert/update supplier items
    task = progress.add_task("[green]Inserting supplier items...", total=len(items))

    for item_agg in items.values():
        supplier_id = supplier_id_map.get(item_agg.supplier_name)
        category_id = category_id_map.get(item_agg.category) if item_agg.category else None

        if not supplier_id:
            progress.advance(task)
            continue

        avg_price = sum(item_agg.prices) / len(item_agg.prices) if item_agg.prices else 0
        min_price = min(item_agg.prices) if item_agg.prices else 0
        max_price = max(item_agg.prices) if item_agg.prices else 0

        existing = session.execute(
            select(SupplierItem).where(
                SupplierItem.supplier_id == supplier_id,
                SupplierItem.item_code == item_agg.item_code,
            )
        ).scalar_one_or_none()

        if existing:
            existing.description = item_agg.description
            existing.supplier_item_code = item_agg.supplier_item_code
            existing.category_id = category_id
            existing.uom = item_agg.uom
            existing.avg_price = avg_price
            existing.min_price = min_price
            existing.max_price = max_price
            existing.total_ordered_qty = item_agg.total_qty
            existing.total_ordered_amount = item_agg.total_amount
            existing.order_count = item_agg.order_count
            existing.last_order_date = item_agg.last_order_date
        else:
            item = SupplierItem(
                supplier_id=supplier_id,
                item_code=item_agg.item_code,
                description=item_agg.description,
                supplier_item_code=item_agg.supplier_item_code,
                category_id=category_id,
                uom=item_agg.uom,
                avg_price=avg_price,
                min_price=min_price,
                max_price=max_price,
                total_ordered_qty=item_agg.total_qty,
                total_ordered_amount=item_agg.total_amount,
                order_count=item_agg.order_count,
                last_order_date=item_agg.last_order_date,
            )
            session.add(item)

        progress.advance(task)

    session.commit()

    # Insert/update supplier-category junctions
    task = progress.add_task(
        "[green]Inserting supplier-category relations...",
        total=len(supplier_categories),
    )

    for sc_agg in supplier_categories.values():
        supplier_id = supplier_id_map.get(sc_agg.supplier_name)
        category_id = category_id_map.get(sc_agg.category_name)

        if not supplier_id or not category_id:
            progress.advance(task)
            continue

        existing = session.execute(
            select(SupplierCategory).where(
                SupplierCategory.supplier_id == supplier_id,
                SupplierCategory.category_id == category_id,
            )
        ).scalar_one_or_none()

        if existing:
            existing.item_count = sc_agg.item_count
            existing.total_amount = sc_agg.total_amount
        else:
            sc = SupplierCategory(
                supplier_id=supplier_id,
                category_id=category_id,
                item_count=sc_agg.item_count,
                total_amount=sc_agg.total_amount,
            )
            session.add(sc)

        progress.advance(task)

    session.commit()


@app.command()
def main(
    excel_path: Path = typer.Option(
        ...,
        "--excel-path",
        "-e",
        help="Path to the Excel file containing PO history data.",
        exists=True,
        dir_okay=False,
        readable=True,
    ),
    db_path: Path = typer.Option(
        Path("data/valerie.db"),
        "--db-path",
        "-d",
        help="Path to the SQLite database file.",
    ),
    drop_existing: bool = typer.Option(
        False,
        "--drop-existing",
        help="Drop existing tables before import.",
    ),
):
    """Import supplier data from Excel file into SQLite database."""
    console.print(f"\n[bold blue]Valerie Chatbot - Excel Data Import[/bold blue]\n")
    console.print(f"Excel file: [green]{excel_path}[/green]")
    console.print(f"Database:   [green]{db_path}[/green]\n")

    # Initialize database
    db = Database(db_path)

    if drop_existing:
        console.print("[yellow]Dropping existing tables...[/yellow]")
        db.drop_tables()

    db.create_tables()

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        TimeElapsedColumn(),
        TimeRemainingColumn(),
        console=console,
    ) as progress:
        # Phase 1: Aggregate data from Excel
        console.print("\n[bold]Phase 1: Reading and aggregating data[/bold]")
        suppliers, items, categories, supplier_categories, legal_entities = aggregate_data(
            excel_path, progress
        )

        console.print(f"\n[cyan]Found:[/cyan]")
        console.print(f"  - {len(suppliers):,} suppliers")
        console.print(f"  - {len(items):,} unique items")
        console.print(f"  - {len(categories):,} categories")
        console.print(f"  - {len(supplier_categories):,} supplier-category relations")
        console.print(f"  - {len(legal_entities):,} legal entities")

        # Phase 2: Insert/update data in database
        console.print("\n[bold]Phase 2: Inserting data into database[/bold]")
        with db.get_session() as session:
            insert_or_update_data(
                session,
                suppliers,
                items,
                categories,
                supplier_categories,
                legal_entities,
                progress,
            )

    console.print(f"\n[bold green]Import complete![/bold green]")
    console.print(f"Database saved to: [green]{db_path.absolute()}[/green]\n")


if __name__ == "__main__":
    app()
